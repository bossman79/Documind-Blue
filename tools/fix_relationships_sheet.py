"""
Normalize relationships.xlsx parent/child flags and add hyperlink column for XlLink-style paths.
Run: python tools/fix_relationships_sheet.py
"""
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "relationships.xlsx"


def is_parent_val(v) -> bool:
    if v is None:
        return False
    s = str(v).strip().lower()
    # Do not treat lone "t" as parent (export used "T" for Is Child only).
    return s in ("y", "yes", "1", "true")


def is_child_val(v) -> bool:
    if v is None:
        return False
    s = str(v).strip().lower()
    return s in ("y", "yes", "1", "true", "t")


def role_label(p: bool, c: bool) -> str:
    if p and c:
        return "Both (parent and child)"
    if p:
        return "Parent only"
    if c:
        return "Child only"
    return "Neither"


def main() -> None:
    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active
    ws.title = "Documents"

    # Expected header row 1
    header = [c.value for c in ws[1]]
    if not header or str(header[0]).strip().lower() != "filename":
        raise SystemExit("Unexpected header row; expected Filename in A1")

    if str(ws.cell(row=1, column=8).value or "").strip() == "Relationship role":
        print(f"{XLSX} already has Relationship role column; not modifying columns.")
        return

    # Insert new columns after G (Is Child): H = Role, I = Hyperlink display+formula
    ws.insert_cols(8, amount=2)
    ws.cell(row=1, column=8, value="Relationship role")
    ws.cell(row=1, column=9, value="File link (uses base path in Z1)")

    z_col = get_column_letter(26)
    ws["Z1"] = ""
    ws["Z2"] = (
        "Optional: set Z1 to a folder that contains these files, with trailing \\ "
        "(e.g. \\\\server\\share\\Drawings and Models\\). "
        "Column I then builds file:/// hyperlinks for Adept XlLink-style discovery."
    )
    ws["Z2"].font = Font(italic=True, size=9)

    max_row = ws.max_row
    for r in range(2, max_row + 1):
        ip = ws.cell(row=r, column=6).value
        ic = ws.cell(row=r, column=7).value
        p, c = is_parent_val(ip), is_child_val(ic)
        ws.cell(row=r, column=6, value="Yes" if p else "No")
        ws.cell(row=r, column=7, value="Yes" if c else "No")
        ws.cell(row=r, column=8, value=role_label(p, c))
        # HYPERLINK: Z1 base + filename in A; empty Z1 leaves link inert
        fn_cell = f"A{r}"
        ws.cell(row=r, column=9, value=f'=IF(${z_col}$1="","",HYPERLINK(${z_col}$1&{fn_cell},{fn_cell}))')

    if "ParentChildEdges" in wb.sheetnames:
        del wb["ParentChildEdges"]
    edges = wb.create_sheet("ParentChildEdges", 1)
    edges.append(
        [
            "Parent Filename",
            "Parent File Id",
            "Child Filename",
            "Child File Id",
            "Notes",
        ]
    )
    edges.append(
        [
            "",
            "",
            "",
            "",
            "This export did not include explicit parent→child pairs. "
            "Fill rows here if you have that mapping (e.g. from Adept or CAD). "
            "You can add HYPERLINK in a fifth sheet or use full paths in a helper column.",
        ]
    )
    edges.column_dimensions["A"].width = 28
    edges.column_dimensions["B"].width = 38
    edges.column_dimensions["C"].width = 28
    edges.column_dimensions["D"].width = 38
    edges.column_dimensions["E"].width = 60

    if "README" in wb.sheetnames:
        del wb["README"]
    readme = wb.create_sheet("README", 0)
    readme.append(["relationships.xlsx — how to use"])
    readme.append([])
    readme.append(
        [
            "Documents!F:G are normalized to Yes/No (was y / 1 / T). "
            "H states whether each file is recorded as a parent, child, or both in your source."
        ]
    )
    readme.append(
        [
            "Set Documents!Z1 to the folder path where these files live (Windows: \\\\server\\share\\folder\\). "
            "Column I builds a clickable link per row for tools that read Excel hyperlinks (e.g. XlLink-style)."
        ]
    )
    readme.append(
        [
            "OLE objects cannot be created by this script; use Insert > Object in Excel if you need OLE links."
        ]
    )
    readme.append(
        [
            "To list explicit parent→child edges, use sheet ParentChildEdges (populate from Adept or xref data)."
        ]
    )
    readme.column_dimensions["A"].width = 100

    wb.save(XLSX)
    print(f"Updated {XLSX}")


if __name__ == "__main__":
    main()
