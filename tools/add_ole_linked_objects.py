"""
Insert linked OLE objects into relationships.xlsx (one per document row).

Requires: Windows, Microsoft Excel installed, pywin32 (`pip install pywin32`).

Example:
  python tools/add_ole_linked_objects.py --base "\\\\fileserver\\library\\Drawings and Models\\"

If Documents!Z1 is set to a folder path (trailing backslash), you can omit --base.

If COM returns "Add method of OLEObjects class failed" (0x800A03EC), Excel/policy is blocking
automated OLE insert. Use tools/AddOLELinks.bas inside Excel instead: Alt+F11, File > Import,
run AddLinkedOLEFromDocuments (or AddFileHyperlinksFromDocuments as fallback).
"""

from __future__ import annotations

import argparse
import os
import sys
import time
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    raise SystemExit(1)

try:
    import win32com.client
except ImportError:
    print("Install pywin32: pip install pywin32", file=sys.stderr)
    raise SystemExit(1)

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WB = ROOT / "relationships.xlsx"


def _normalize_base(s: str) -> str:
    s = (s or "").strip()
    if not s:
        return ""
    return s if s.endswith(("\\", "/")) else s + "\\"


def read_filenames_and_z1(xlsx: Path) -> tuple[str, list[str]]:
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    if "Documents" not in wb.sheetnames:
        raise SystemExit("Workbook must contain a 'Documents' sheet.")
    doc = wb["Documents"]
    z1 = _normalize_base(str(doc["Z1"].value or ""))
    names: list[str] = []
    for row in range(2, doc.max_row + 1):
        v = doc.cell(row=row, column=1).value
        if v is None or str(v).strip() == "":
            continue
        names.append(str(v).strip())
    return z1, names


def ensure_ole_sheet(wb_com, after_name: str = "Documents"):
    for sh in wb_com.Worksheets:
        if sh.Name == "OLE_Links":
            return sh
    sh = wb_com.Worksheets.Add(After=wb_com.Worksheets(after_name))
    sh.Name = "OLE_Links"
    sh.Range("A1").Value = "Filename"
    sh.Range("B1").Value = "Linked OLE (icon)"
    sh.Range("C1").Value = "Status"
    return sh


def delete_all_ole(ws):
    oles = ws.OLEObjects()
    n = int(oles.Count)
    for i in range(n, 0, -1):
        oles.Item(i).Delete()


def main() -> None:
    ap = argparse.ArgumentParser(description="Add linked OLE objects for each file in relationships.xlsx")
    ap.add_argument("--workbook", type=Path, default=DEFAULT_WB, help="Path to relationships.xlsx")
    ap.add_argument(
        "--base",
        type=str,
        default="",
        help="Folder containing the files (trailing \\ optional). Overrides Documents!Z1 if set.",
    )
    ap.add_argument("--visible", action="store_true", help="Show Excel while running (debug).")
    args = ap.parse_args()

    xlsx = args.workbook.resolve()
    if not xlsx.is_file():
        raise SystemExit(f"Workbook not found: {xlsx}")

    z1, filenames = read_filenames_and_z1(xlsx)
    base = _normalize_base(args.base) if args.base.strip() else z1
    if not base:
        raise SystemExit(
            "Set Documents!Z1 to the files folder (with trailing backslash) or pass --base \"\\\\server\\share\\folder\\\""
        )

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = bool(args.visible)
    excel.DisplayAlerts = False
    wb = None
    try:
        wb = excel.Workbooks.Open(str(xlsx))
        doc_ws = wb.Worksheets("Documents")
        ole_ws = ensure_ole_sheet(wb)
        delete_all_ole(ole_ws)

        row = 2
        ok = 0
        root = base.rstrip("/\\")
        for fn in filenames:
            full_str = os.path.normpath(os.path.join(root, fn))
            ole_ws.Cells(row, 1).Value = fn

            if not os.path.isfile(full_str):
                ole_ws.Cells(row, 3).Value = f"Skipped (not found): {full_str}"
                row += 1
                continue

            left = float(ole_ws.Cells(row, 2).Left)
            top = float(ole_ws.Cells(row, 2).Top)
            width, height = 72.0, 72.0

            try:
                # Linked package to file — what Adept “OLE relationships” refers to.
                ole_ws.OLEObjects().Add(
                    Filename=full_str,
                    Link=True,
                    DisplayAsIcon=True,
                    Left=left,
                    Top=top,
                    Width=width,
                    Height=height,
                )
                ole_ws.Cells(row, 3).Value = "OK (linked OLE)"
                ok += 1
            except Exception as e:  # noqa: BLE001
                ole_ws.Cells(row, 3).Value = f"Error: {e!s}"

            row += 1
            time.sleep(1.5)

        wb.Save()
        print(f"Done. Linked OLE inserts attempted: {ok} OK of {len(filenames)} rows. See sheet OLE_Links column C.")
    finally:
        if wb is not None:
            wb.Close(SaveChanges=True)
        excel.Quit()


if __name__ == "__main__":
    main()
