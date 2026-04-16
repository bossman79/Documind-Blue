#!/usr/bin/env python3
"""
Read relationships.xlsx (Sheet1, column A = Filename), pair same-stem .pdf + .dwg,
and write linked PDFs.

Default (no file paths needed): each output PDF uses file:./ChildName.dwg links — the
DWG is resolved next to that PDF in the same folder (co-located), matching typical
xref-style “same directory” behavior without UNC columns.

Optional: --base if you have source PDFs on disk to start from; --absolute-paths for
full file:// URIs (requires files under --base).

Requires: pip install pymupdf openpyxl

Examples:
  python tools/run_relationships_pdf_links.py -o .\\linked_out\\
  python tools/run_relationships_pdf_links.py --from-folder C:\\pack\\ -o .\\linked_out\\
  python tools/run_relationships_pdf_links.py --base \\\\server\\share\\folder\\ -o .\\out\\
  python tools/run_relationships_pdf_links.py --absolute-paths --base C:\\drawings\\ -o .\\out\\
"""

from __future__ import annotations

import argparse
import sys
from collections.abc import Callable
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

try:
    import fitz
except ImportError:
    print("Install PyMuPDF: pip install pymupdf", file=sys.stderr)
    sys.exit(1)

_TOOLS = Path(__file__).resolve().parent
if str(_TOOLS) not in sys.path:
    sys.path.insert(0, str(_TOOLS))

from pdf_set_child_file_links import (  # noqa: E402
    add_child_file_links,
    add_child_links_by_filename,
)


def discover_pairs_in_folder(folder: Path) -> list[tuple[str, Path, Path]]:
    """Same stem .pdf + .dwg in one folder → (stem, pdf_path, dwg_path)."""
    folder = folder.resolve()
    pdfs = {p.stem.lower(): p for p in folder.glob("*.pdf")}
    dwgs = {p.stem.lower(): p for p in folder.glob("*.dwg")}
    pairs: list[tuple[str, Path, Path]] = []
    for stem_key in sorted(set(pdfs) & set(dwgs), key=lambda s: pdfs[s].stem):
        pairs.append((pdfs[stem_key].stem, pdfs[stem_key], dwgs[stem_key]))
    return pairs


def load_stem_pairs(xlsx: Path) -> list[tuple[str, str, str]]:
    """Return list of (stem, pdf_filename, dwg_filename)."""
    ws = openpyxl.load_workbook(xlsx, data_only=True).active
    by_stem: dict[str, dict[str, str]] = {}
    for r in range(2, ws.max_row + 1):
        fn = ws.cell(r, 1).value
        if not fn or not str(fn).strip():
            continue
        fn = str(fn).strip()
        p = Path(fn)
        suf = p.suffix.lower()
        if suf not in (".pdf", ".dwg"):
            continue
        by_stem.setdefault(p.stem, {})[suf] = fn
    pairs: list[tuple[str, str, str]] = []
    for stem, parts in sorted(by_stem.items()):
        if ".pdf" in parts and ".dwg" in parts:
            pairs.append((stem, parts[".pdf"], parts[".dwg"]))
    return pairs


def bootstrap_stub_files(xlsx: Path, base: Path) -> None:
    """Create minimal PDFs and empty DWGs for every filename in column A."""
    ws = openpyxl.load_workbook(xlsx, data_only=True).active
    base.mkdir(parents=True, exist_ok=True)
    seen: set[str] = set()
    for r in range(2, ws.max_row + 1):
        fn = ws.cell(r, 1).value
        if not fn or not str(fn).strip():
            continue
        fn = str(fn).strip()
        if fn in seen:
            continue
        seen.add(fn)
        dest = base / fn
        dest.parent.mkdir(parents=True, exist_ok=True)
        lower = fn.lower()
        if lower.endswith(".pdf"):
            doc = fitz.open()
            doc.new_page()
            doc.save(dest)
            doc.close()
        else:
            dest.touch()


def run_excel_batch(
    xlsx: Path,
    out_dir: Path,
    *,
    base: Path | None = None,
    absolute_paths: bool = False,
    bootstrap_stubs: bool = False,
    labels: bool = True,
    log: Callable[[str], None] | None = None,
) -> tuple[int, int]:
    """Returns (ok_count, skipped_count)."""
    lg = log or (lambda m: print(m, file=sys.stderr))

    if bootstrap_stubs:
        if base is None:
            raise ValueError("--bootstrap-stubs requires --base")
        bootstrap_stub_files(xlsx, base)
        lg(f"Bootstrapped stub files under {base}")

    pairs = load_stem_pairs(xlsx)
    if not pairs:
        raise ValueError("No stem pairs with both .pdf and .dwg in spreadsheet.")

    out_dir.mkdir(parents=True, exist_ok=True)
    ok = skipped = 0
    for stem, pdf_name, dwg_name in pairs:
        out_pdf = out_dir / f"{stem}_linked.pdf"
        try:
            if absolute_paths:
                if base is None:
                    raise ValueError("absolute_paths requires base folder")
                pdf_path = base / pdf_name
                dwg_path = base / dwg_name
                if not pdf_path.is_file() or not dwg_path.is_file():
                    lg(f"Skip {stem}: missing files under {base}")
                    skipped += 1
                    continue
                add_child_file_links(
                    pdf_path,
                    [dwg_path],
                    out_pdf,
                    labels=labels,
                )
            else:
                src_pdf = (base / pdf_name) if base is not None else None
                parent = src_pdf if src_pdf is not None and src_pdf.is_file() else None
                add_child_links_by_filename(
                    out_pdf,
                    [dwg_name],
                    parent_pdf=parent,
                    labels=labels,
                )
            ok += 1
            lg(str(out_pdf))
        except Exception as e:  # noqa: BLE001
            lg(f"Error {stem}: {e}")
            skipped += 1
    lg(f"Done: {ok} linked, {skipped} skipped, {len(pairs)} pairs in sheet.")
    return ok, skipped


def run_folder_batch(
    input_dir: Path,
    out_dir: Path,
    *,
    absolute_paths: bool = False,
    labels: bool = True,
    log: Callable[[str], None] | None = None,
) -> tuple[int, int]:
    """Pair PDF+DWG by stem in input_dir; write *_linked.pdf to out_dir."""
    lg = log or (lambda m: print(m, file=sys.stderr))
    pairs = discover_pairs_in_folder(input_dir)
    if not pairs:
        raise ValueError("No matching .pdf + .dwg pairs (same base name) in folder.")

    out_dir.mkdir(parents=True, exist_ok=True)
    ok = skipped = 0
    for stem, pdf_path, dwg_path in pairs:
        out_pdf = out_dir / f"{stem}_linked.pdf"
        try:
            if absolute_paths:
                add_child_file_links(
                    pdf_path,
                    [dwg_path],
                    out_pdf,
                    labels=labels,
                )
            else:
                add_child_links_by_filename(
                    out_pdf,
                    [dwg_path.name],
                    parent_pdf=pdf_path,
                    labels=labels,
                )
            ok += 1
            lg(str(out_pdf))
        except Exception as e:  # noqa: BLE001
            lg(f"Error {stem}: {e}")
            skipped += 1
    lg(f"Done: {ok} linked, {skipped} skipped, {len(pairs)} pairs in folder.")
    return ok, skipped


def main() -> None:
    root = Path(__file__).resolve().parents[1]
    ap = argparse.ArgumentParser(
        description="Pair PDF+DWG by stem from relationships.xlsx; add DWG links (relative names by default)."
    )
    ap.add_argument(
        "--relationships",
        type=Path,
        default=root / "relationships.xlsx",
        help="Path to relationships.xlsx",
    )
    ap.add_argument(
        "--base",
        type=Path,
        default=None,
        metavar="SOURCE_PDF_FOLDER",
        help="Folder containing source .pdf files (optional). If set and the PDF exists, it is used as the parent; else a blank PDF is used.",
    )
    ap.add_argument(
        "--from-folder",
        type=Path,
        default=None,
        help="Pair *.pdf + *.dwg with same base name in this folder (ignore Excel)",
    )
    ap.add_argument(
        "-o",
        "--output-dir",
        type=Path,
        required=True,
        help="Folder for output *_linked.pdf files",
    )
    ap.add_argument(
        "--absolute-paths",
        action="store_true",
        help="Use full file:// URIs; requires --base and both PDF+DWG files present there",
    )
    ap.add_argument(
        "--bootstrap-stubs",
        action="store_true",
        help="Create minimal PDFs and empty DWGs under --base from the sheet (testing only)",
    )
    ap.add_argument("--no-labels", action="store_true")
    args = ap.parse_args()

    out_dir = args.output_dir.resolve()

    if args.from_folder is not None:
        inp = args.from_folder.resolve()
        if not inp.is_dir():
            sys.exit(f"Not a folder: {inp}")
        try:
            run_folder_batch(
                inp,
                out_dir,
                absolute_paths=args.absolute_paths,
                labels=not args.no_labels,
                log=lambda m: print(m, file=sys.stderr),
            )
        except ValueError as e:
            sys.exit(str(e))
        return

    xlsx = args.relationships.resolve()
    if not xlsx.is_file():
        sys.exit(f"Spreadsheet not found: {xlsx}")

    if args.absolute_paths and args.base is None:
        sys.exit("--absolute-paths requires --base")

    if args.bootstrap_stubs and args.base is None:
        sys.exit("--bootstrap-stubs requires --base")

    base = args.base.resolve() if args.base else None

    try:
        run_excel_batch(
            xlsx,
            out_dir,
            base=base,
            absolute_paths=args.absolute_paths,
            bootstrap_stubs=bool(args.bootstrap_stubs and base is not None),
            labels=not args.no_labels,
            log=lambda m: print(m, file=sys.stderr),
        )
    except ValueError as e:
        sys.exit(str(e))


if __name__ == "__main__":
    main()
